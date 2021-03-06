#!/usr/bin/env python
from cached_property import cached_property
from openpyxl import load_workbook

from EPPs.common import StepEPP, InvalidStepError, RestCommunicationEPP


class ParseManifest(StepEPP, RestCommunicationEPP):
    """
    Scripts for parsing sample manifest for either plates, tubes or SGP tubes. The key used for the parsing is the Sample
    Name for plates and the 2D barcode for tubes/SGP tubes. The script assembles a dictionary of the sample root artifacts
    using the appropriate keys. It obtains all the sample UDFs to be parsed by finding any step UDFs with the tag [Plate]
    [Tube] or [SGP]. If the sample UDFs that should be parsed needs to changed then the manifest document should be
    changed and the step UDFs updated accordingly. The step UDFs contain the column in the manifest that contains that
    UDF. The step UDF name must match the sample UDF name. Please note that assignment of UDFs in the LIMS as either
    numeric (int) or single-line text (str) is not always logical so the [Str] tag was introduced to the step UDF naming
    to convert problem data points to strings where required.
    """

    # additional argument required to obtain the file location for newly create manifest in the LIMS step
    def __init__(self, argv=None):
        super().__init__(argv)
        self.manifest = self.cmd_args.manifest

    @staticmethod
    def add_args(argparser):
        argparser.add_argument('-m', '--manifest', type=str, required=True, help='Sample manifest returned by customer')

    def _validate_step(self):
        super()._validate_step()
        # Check that only one container type is present in step
        unique_container_types = set([artifact.container.type for artifact in self.artifacts])
        if len(unique_container_types) > 1:
            raise InvalidStepError(message='Multiple container types present in step. Only 1 container type permitted')

    def validate_species(self, species_name):
        """Validate species name against REST API."""
        if self.get_documents('species', where={'name': species_name}):
            return True
        return False

    def validate_genome_version(self, genome_version):
        """Validate genome version against REST API unless it is empty."""
        if genome_version == '' or self.get_documents('genomes', where={'assembly_name': genome_version}):
            return True
        return False

    def setup_parsing_parameters(self):
        """
        Identify the container type present in the step and assign variables appropriately. con_type is the step UDF
        tag, key is the variable that is used for linking the sample to the matching manifest row. key column is the
        column in the manifest that holds the key information and current_row starts as the first row in the excel
        sheet that contains sample data and is then updated as the manifest is parsed
        """
        container_type = self.process.all_inputs()[0].container.type.name
        if container_type == '96 well plate':
            self.con_type = '[Plate]'
            self.sample_column_udf = 'Plate Sample Name'
            self.sample_column = self.process.udf[self.sample_column_udf]
            self.current_row = self.process.udf['Plate Starting Row']

        elif container_type == 'rack 96 positions':
            self.con_type = '[Tube]'
            self.sample_column_udf = '2D Barcode'
            self.sample_column = self.process.udf[self.sample_column_udf]
            self.current_row = self.process.udf['Tube Starting Row']

        elif container_type == 'SGP rack 96 positions':
            self.con_type = '[SGP]'
            self.sample_column_udf = '2D Barcode'
            self.sample_column = self.process.udf[self.sample_column_udf]
            self.current_row = self.process.udf['SGP Starting Row']

    @cached_property
    def step_udfs_to_parse(self):
        """identify the udfs that should be parsed for this container type"""
        return list(set(udf_name for udf_name in self.process.udf if self.con_type in udf_name))

    def _run(self):
        self.setup_parsing_parameters()

        manifest_file = self.open_or_download_file(self.manifest, binary=True)
        # open the excel manifest
        wb = load_workbook(manifest_file, data_only=True)
        ws = wb.active
        # create the dictionary of sample based on the appropriate key (sample id or 2D barcode) for the container type
        sample_dict = {}
        for sample in self.samples:
            if self.sample_column_udf == 'Plate Sample Name':
                sample_dict[sample.name] = sample
            if self.sample_column_udf == '2D Barcode':
                sample_dict[sample.udf['2D Barcode']] = sample

        # create the list to contain the samples to be updated in a batch PUT to the API. This should match the number
        # of input artifacts
        samples_to_put = []

        sample_id = ws[self.sample_column + str(self.current_row)].value
        while sample_id:
            if sample_id not in sample_dict:
                # Raise value error if sample present in manifest but not in LIMS
                raise InvalidStepError(sample_id + " present in manifest but not present in LIMS.")

            for udf in self.step_udfs_to_parse:
                # remove the container type tag from the step UDF name so can be used to find the corresponding sample UDF
                lims_udf = udf.replace(self.con_type, '')
                # use the column indicated in the step UDF and the current row counter to identify which cell in
                # the spreadsheet should be referenced
                udf_cell = self.process.udf[udf] + str(self.current_row)
                # check for udfs that need to be converted to strings as indicated by the [Str] tag in the step UDF name

                if lims_udf.find('[Str]') >= 0:
                    lims_udf = lims_udf.replace('[Str]', '')
                    udf_value = str(ws[udf_cell].value)
                else:
                    udf_value = ws[udf_cell].value

                if udf_value is None:
                    self.warning('Value UDF %s is not set for sample %s (cell %s).', lims_udf, sample_id, udf_cell)
                    continue

                if 'Species' in lims_udf and not self.validate_species(udf_value):
                    raise InvalidStepError(udf_value + ' in ' + lims_udf + ' is not a valid species')
                elif 'Genome Version' in lims_udf and not self.validate_genome_version(udf_value):
                    raise InvalidStepError(udf_value + ' in ' + lims_udf + ' is not a valid genome version')

                # parse the data from the spreadsheet into the sample dictionary
                sample_dict[sample_id].udf[lims_udf] = udf_value

            # add the modified sample to the list of modified samples to be put into the database.
            samples_to_put.append(sample_dict[sample_id])

            self.current_row += 1
            sample_id = ws[self.sample_column + str(self.current_row)].value

        # check that there not fewer samples in the manifest than in the LIMS
        if not len(samples_to_put) == len(self.process.all_inputs(unique=True)):
            raise InvalidStepError('The number of samples in the step does not match the number of samples in the manifest')

        self.lims.put_batch(samples_to_put)


if __name__ == '__main__':
    ParseManifest().run()
