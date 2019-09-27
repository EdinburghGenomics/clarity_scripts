#!/usr/bin/env python
from openpyxl.styles import Font

from EPPs.common import SendMailEPP
from datetime import date
from openpyxl import Workbook


# batch limit function reduces the number of samples in each batch call to prevent server resource error (modified from prepare_discard_plate.py)
def batch_limit(lims, list_instance, max_query=100):
    submitted_samples_batch = []
    for start in range(0, len(list_instance), max_query):
        submitted_samples_batch = submitted_samples_batch + lims.get_batch(list_instance[start:start + max_query])
    return submitted_samples_batch


def get_human_artifacts(lims):
    # find all the human submitted samples in the LIMS
    submitted_samples = lims.get_samples(udf={'Species': 'human'}) + lims.get_samples(
        udf={'Species': 'Homo sapiens'})
    # it is too slow querying each sample for udf and project details so make batch calls to the LIMS using batch_limit function
    submitted_samples_batch = batch_limit(lims, submitted_samples)
    return submitted_samples_batch


def create_excel_report(lims, report_path):
    # obtain the samples present in the LIMS that are from humans and not showing as "Sample Destroyed"
    submitted_samples_batch = get_human_artifacts(lims)

    # create a new xlsx workbook and obtain the active worksheet that is created by default
    wb = Workbook()
    ws = wb.active
    ws.title = 'Human Tissue Report'
    # define lists containing the column headers, the excel columns and the widths of the excel columns
    column_headers = ['PI', 'Sample Type', 'Project Name', 'Submitted Sample Name', 'REC#', 'Ethics Committee',
                      'Freezer', 'Shelf']
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    widths = [20, 15, 15, 25, 20, 60, 20, 20]

    # populate the active worksheet with the column headers and widths as well as formatting fonr to bold
    for header, column, width in zip(column_headers, columns, widths):
        ws[column + '1'] = header
        ws[column + '1'].font = Font(bold=True)
        ws.column_dimensions[column].width = width

    # populate the report with the sample information
    # start the row counter at 2 as the headings are in row 1
    row_counter = 2

    # create a dictionary where key is the sample name so can sort later by reverse sample name when writing csv
    samples_dict = {}
    for sample in submitted_samples_batch:
        samples_dict[sample.name] = sample

    sample_keys_list = samples_dict.keys()

    # sort sample keys
    sample_keys_list.sort(reverse=True)

    # write excel file in sample name reverse alphanumerical order
    for key in sample_keys_list:
        freezer_location = samples_dict[key].udf.get('Freezer')
        project_name = samples_dict[key].project.name
        # only include samples where the freezer location is not "Sample Destroyed" and the project is an actual
        # project identified by starting with an X
        if freezer_location != 'Sample Destroyed' and project_name[0] == 'X1':
            shelf = samples_dict[key].udf.get('Shelf')
            ws['A' + str(row_counter)] = 'Edinburgh Genomics'
            ws['B' + str(row_counter)] = 'DNA'
            ws['C' + str(row_counter)] = project_name
            ws['D' + str(row_counter)] = samples_dict[key].name
            ws['E' + str(row_counter)] = samples_dict[key].project.udf.get('REC#')
            ws['F' + str(row_counter)] = samples_dict[key].project.udf.get(
                'Organisation providing ethical consent approval')
            ws['G' + str(row_counter)] = freezer_location
            ws['H' + str(row_counter)] = shelf
            row_counter += 1

    wb.save(filename=report_path)


class GenerateHumanTissueReport(SendMailEPP):
    """Generate an excel file containing the details of samples in LIMS where the sample UDF Species is human or homo sapiens and the freezer location
    is not Sample Destroyed"""

    # today's date required for file name and email
    todays_date = str(date.today())
    # csv report needs to be accessible throughout object
    report_path = 'Human_Tissue_Report_' + todays_date + '.xlsx'

    def send_email(self):
        msg_subject = "Human Tissue Report - Edinburgh Genomics - " + self.todays_date
        msg = "Hi,\n\nPlease find attached the Human Tissue Report from Edinburgh Genomics Clinical for %s.\n\nKind Regards,\nEdinburgh Genomics" % self.todays_date
        self.send_mail(msg_subject, msg, attachments=[self.report_path], config_name='human_tissue')

    def _run(self):
        create_excel_report(self.lims, self.report_path)
        self.send_email()


if __name__ == '__main__':
    GenerateHumanTissueReport().run()
