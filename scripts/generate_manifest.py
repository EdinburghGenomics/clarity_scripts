#!/usr/bin/env python
import itertools
from openpyxl import load_workbook
import os


from EPPs.common import SendMailEPP


class GenerateManifest96WellPlate(SendMailEPP):
    # populate the sample manifest with the sample date. Sample manifest template is determined by a step udf.
    # The starting row and columns are determined by step UDFs.

    _use_load_config = True  # prevent the loading of the config file

    # additional argument required to obtain the file location for newly create manifest in the LIMS step
    def __init__(self, argv=None):
        super().__init__(argv)
        self.manifest = self.cmd_args.manifest

    @staticmethod
    def add_args(argparser):
        argparser.add_argument(
            '-m', '--manifest', type=str, required=True, help='Sample manifest generated by the LIMS'
        )

    def _run(self):

        # obtain all of the inputs for the step
        all_inputs = self.process.all_inputs(unique=True)

        # check all input containers have the same type and also create list of unique containers
        container_types = set()
        unique_containers=set()


        for artifact in all_inputs:
            container_types.add(artifact.container.type)
            unique_containers.add(artifact.container)

        if len(container_types) > 1:
            raise ValueError('Only 1 container type is permitted. Multiple container types are present in the step')

        # obtain step udfs
        step_udfs = self.process.udf

        # obtain the name of container type of the samples
        if list(container_types)[0].name == '96 well plate':
            con_type = '[Plates]'
        elif list(container_types)[0].name == 'rack 96 positions' and self.process.udf['Are there special project requirements?']=='No':
            con_type = '[Tubes]'
        elif list(container_types)[0].name == 'rack 96 positions' and self.process.udf['Are there special project requirements?']=='Scottish Genome Partnership':
            con_type = '[SGP]'

        # define counter to ensure each sample is written to a new well
        row_counter = step_udfs[con_type + 'Starting Row']

        # open the correct manifest template for the container type
        wb = load_workbook(filename=step_udfs[con_type + 'Manifest File Path'])
        ws = wb.active

        configurable_udfs = []

        for step_udf_key in step_udfs.keys():
            tag_position = step_udf_key.find('[Sample UDF]')

            if tag_position > -1:
                configurable_udfs.append(step_udf_key[tag_position + 12:])

        # define the rows and columns of the 96 well plate/rack to be used for writing the manifest in correct order
        rows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        columns = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']

        sample_dict={}

        # create a dictionary of samples so the manifest can be written by column
        for artifact in all_inputs:
            sample_dict[artifact.location[1].replace(":","")]=artifact.samples[0]


        for column,row in itertools.product(columns,rows):
            if row+column not in sample_dict:
                continue
            #if condition met then 'continue' means will skip back to beginning of loop

            # populate the manifest with sample attributes
            sample_udf=sample_dict[row+column].udf
            #sample_udf = artifact.samples[0].udf
            # populate the wells in the active work sheet in the excel in the columns defined by the step UDFs
            if con_type == '[Plates]':
                ws[step_udfs[con_type + 'Sample Name'] + str(row_counter)] = sample_dict[row+column].artifact.name
                ws[step_udfs['[Plates]Container Name'] + str(row_counter)] = sample_dict[row+column].artifact.container.name
                ws[step_udfs['[Plates]Well'] + str(row_counter)] = sample_dict[row+column].artifact.location[1]
                ws[step_udfs['[Plates]Project ID'] + str(row_counter)] =sample_dict[row+column].project.name

            if con_type == '[Tubes]':
                ws[step_udfs['[Tubes]Project ID Well']] = sample_dict[row+column].project.name


            if con_type == '[SGP]':
                ws[step_udfs['[SGP]Project ID Well']] = sample_dict[row+column].project.name


            # populate the manifest with sample UDFs which are configurable by adding or removing step UDFs in the
            # format [CONTAINER TYPE - either Tubes or Plates][Sample UDF]Name of UDF
            for configurable_udf in configurable_udfs:
                if con_type + '[Sample UDF]' + configurable_udf in step_udfs.keys():
                    ws[step_udfs[con_type + '[Sample UDF]' + configurable_udf] + str(row_counter)] = \
                        sample_udf[configurable_udf]

            row_counter += 1



        # create a new file with the original file name plus a suffix containing the project ID
        lims_filepath = self.manifest + '-'+'Edinburgh_Genomics_Sample_Submission_Manifest_' + all_inputs[0].samples[0].project.name + '.xlsx'
        wb.save(filename=lims_filepath)
        #write a copy of the file without the LIMS prefix that can be attached to a customer email
        email_filepath='Edinburgh_Genomics_Sample_Submission_Manifest_' + all_inputs[0].samples[0].project.name + '.xlsx'
        wb = load_workbook(filename=lims_filepath)
        wb.save(filename=email_filepath)

        #send an email to the project manager using customer manifest template with the new manifest attached
        project_name = all_inputs[0].samples[0].project.name
        #choose a species for the email subject, this can be updated manually by the project manager
        species=all_inputs[0].samples[0].udf['Species']
        email_subject= project_name+": "+species+" WGS Sample Submission"

        #the manifest and relevant requirements document for the container type should be attached to the email.
        attachments_list=[]
        attachments_list.append(email_filepath)
        if list(container_types)[0].name=='96 well plate':
            attachments_list.append(self.process.udf['Plates Requirements Path'])
        elif list(container_types)[0].name=='rack 96 positions':
            attachments_list.append(self.process.udf['Tubes Requirements Path'])

        self.send_mail(email_subject,None, project=project_name,template_name='customer_manifest.html',
                       config_name='projects-facility-finance_only', attachments=attachments_list)
        #delete the copy of the manifest that was attached to the email
        os.remove(email_filepath)



if __name__ == '__main__':
    GenerateManifest96WellPlate().run()
